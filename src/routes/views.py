import io
import openpyxl
from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, current_app, Response
from forms import BoatRegistrationForm, StatusCheckForm, BoatEditForm
from db import add_boat_instance, get_all_boats, delete_boat, get_boat_by_id, update_boat
from services.reservation_checker import check_single_boat
from forms import REGION_CHOICES
from datetime import date as dt_date
from urllib.parse import urlparse
from models import Boat
from concurrent.futures import ThreadPoolExecutor, as_completed
import re

views = Blueprint('views', __name__, template_folder='templates')

@views.route('/')
def index():
    boats = get_all_boats()
    # Boat ê°ì²´ë“¤ì„ ë”•ì…”ë„ˆë¦¬ë¡œ ë³€í™˜í•˜ì—¬ JSON ì§ë ¬í™” ê°€ëŠ¥í•˜ê²Œ ë§Œë“­ë‹ˆë‹¤
    boats_dict = [boat.to_dict() for boat in boats]

    # í™ˆ ëª¨ë‹¬ ë“±ë¡ í¼ì—ì„œ CSRF ë¥¼ ì‚¬ìš©í•˜ê¸° ìœ„í•´ í¼ ì¸ìŠ¤í„´ìŠ¤ë¥¼ ì „ë‹¬
    form = BoatRegistrationForm()

    return render_template(
        'index.html',
        boats=boats,
        boats_json=boats_dict,
        form=form,
        city_port_map=city_port_mapping
    )

@views.route('/download_excel')
def download_excel():
    boats = get_all_boats()
    
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registered Boats"
    
    # Add header row
    headers = ["No", "ì§€ì—­", "í•­êµ¬", "ë“±ë¡ëœ ë°°", "URL"]
    ws.append(headers)
    
    # Add data rows
    for i, boat in enumerate(boats, start=1):
        row = [i, boat.city, boat.port, boat.name, boat.url]
        ws.append(row)
        
    # Create a virtual file to save the workbook
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create a response
    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=boat_list.xlsx"}
    )

city_port_mapping = {
    'ì¸ì²œ': ['ë‚¨í•­(ì¸ì²œí•­)', 'ì—°ì•ˆë¶€ë‘', 'ì˜í¥í•­'],
    'ì•ˆì‚°': ['ì˜¤ì´ë„í•­'],
    'í™”ì„±': ['ì „ê³¡í•­'],
    'í‰íƒ': ['í‰íƒí•­'],
    'ë‹¹ì§„': ['ì¥ê³ í•­'],
    'ì„œì‚°': ['ì‚¼ê¸¸í¬í•­'],
    'íƒœì•ˆ': ['ë§ˆê²€í¬í•­', 'ëª¨í•­í•­', 'ì˜ëª©í•­', 'ì‹ ì§„ë„í•­'],
    'ë³´ë ¹': ['ì˜¤ì²œí•­', 'êµ¬ë§¤í•­', 'ëŒ€ì²œí•­', 'ë¬´ì°½í¬í•­', 'ë‚¨ë‹¹í•­', 'í™ì›í•­'],
    'êµ°ì‚°': ['ë¹„ì‘í•­', 'ì•¼ë¯¸ë„í•­'],
    'ê²©í¬': ['ê²©í¬í•­'],
    'ì—¬ìˆ˜': ['ëŒì‚°í•­', 'êµ­ë™í•­', 'ì†Œí˜¸í•­', 'ì‹ ì¶”í•­', 'ì¢…í¬í•­'],
    'ê³ í¥': ['ë…¹ë™ë°©íŒŒì œ']
}

@views.route('/register', methods=['GET', 'POST'])
def register():
    form = BoatRegistrationForm()
    if request.method == 'POST':
        city = request.form.get('city')
        if city in city_port_mapping:
            form.port.choices = [(port, port) for port in city_port_mapping[city]]
    
    if form.validate_on_submit():
        try:
            add_boat_instance(form.name.data, form.url.data, form.city.data, form.port.data, form.note.data)
            flash('ë°°ê°€ ë“±ë¡ë˜ì—ˆìŠµë‹ˆë‹¤.', 'success')
            return redirect(url_for('views.index'))
        except Exception as e:
            flash(f'ë“±ë¡ ì¤‘ ì˜¤ë¥˜: {e}', 'danger')
    return render_template('register.html', form=form)

@views.route('/edit/<int:boat_id>', methods=['GET', 'POST'])
def edit_boat(boat_id):
    boat = get_boat_by_id(boat_id)
    if not boat:
        flash('í•´ë‹¹ ë°°ë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤.', 'danger')
        return redirect(url_for('views.index'))

    form = BoatEditForm(obj=boat)
    if request.method == 'POST':
        city = request.form.get('city')
        if city in city_port_mapping:
            form.port.choices = [(port, port) for port in city_port_mapping[city]]

        if form.validate_on_submit():
            try:
                update_boat(boat_id, form.name.data, form.url.data, form.city.data, form.port.data, form.note.data)
                flash('ë°° ì •ë³´ê°€ ìˆ˜ì •ë˜ì—ˆìŠµë‹ˆë‹¤.', 'success')
                return redirect(url_for('views.index'))
            except Exception as e:
                flash(f'ìˆ˜ì • ì¤‘ ì˜¤ë¥˜: {e}', 'danger')
    else:
        # GET ìš”ì²­ ì‹œ, í˜„ì¬ ë„ì‹œì˜ í•­êµ¬ ëª©ë¡ì„ ì„¤ì •
        if boat.city in city_port_mapping:
            form.port.choices = [(port, port) for port in city_port_mapping[boat.city]]

    return render_template('edit_boat.html', form=form, boat_id=boat_id)

@views.route('/status', methods=['GET'])
def status():
    form = StatusCheckForm()
    # ì¿¼ë¦¬ì—ì„œ ê°’ ì½ì–´ í¼ì— ì£¼ì… (ì¡°íšŒ í›„ì—ë„ ê°’ ìœ ì§€)
    y_arg = request.args.get("year")
    m_arg = request.args.get("month")
    d_arg = request.args.get("day")
    if y_arg: form.year.data = int(y_arg)
    if m_arg: form.month.data = int(m_arg)
    if d_arg: form.day.data = int(d_arg)

    # ì§€ì—­ ëª©ë¡ ë° ì„ íƒê°’
    region_names = [label for value, label in REGION_CHOICES if value]
    selected_regions = request.args.getlist("regions") or ['ì „ì²´']

    # --- added: compute region_counts immediately so status page shows counts on load ---
    registered_boats = get_all_boats()
    region_sets = {}
    for b in registered_boats:
        city = getattr(b, 'city', None) or ''
        rn = getattr(b, 'name', None) or getattr(b, 'registered_name', None) or ''
        if not city:
            continue
        region_sets.setdefault(city, set()).add(rn or '')
    region_counts = { r: len(s) for r, s in region_sets.items() }
    total_registered = sum(region_counts.values())
    # --- end added ---

    # ë‚ ì§œ ë¯¸ì…ë ¥ ì‹œ ì¡°íšŒí•˜ì§€ ì•Šê³  í™”ë©´ë§Œ ë Œë”ë§
    if not (y_arg and m_arg and d_arg):
        return render_template(
            "status.html",
            form=form,
            entries=[],
            year=y_arg or "",
            month=m_arg or "",
            day=d_arg or "",
            region_names=region_names,
            selected_regions=selected_regions,
            region_counts=region_counts,        # now populated
            total_registered=total_registered   # now populated
        )

    # ë‚ ì§œ íŒŒì‹±
    try:
        year, month, day = int(y_arg), int(m_arg), int(d_arg)
    except Exception:
        flash("ì—°/ì›”/ì¼ì„ ì˜¬ë°”ë¥´ê²Œ ì…ë ¥í•˜ì„¸ìš”.", "warning")
        return render_template(
            "status.html",
            form=form,
            entries=[],
            year=y_arg or "",
            month=m_arg or "",
            day=d_arg or "",
            region_names=region_names,
            selected_regions=selected_regions,
            region_counts={},           # { changed code }
            total_registered=0          # { changed code }
        )

    # ì§€ì—­ í•„í„°ë§(OR). 'ì „ì²´'ë§Œ ì„ íƒ ì‹œ ì „ì²´ ì¡°íšŒ
    registered_boats = get_all_boats()

    # { changed code } : ì„ íƒëœ ì§€ì—­ì— ë”°ë¼ ì¿¼ë¦¬ ëŒ€ìƒ ëª©ë¡ ìƒì„±
    filter_targets = [r for r in selected_regions if r != 'ì „ì²´']
    boats_to_query = [b for b in registered_boats if b.city in filter_targets] if filter_targets else registered_boats

    # DEBUG: get_all_boats() ë°˜í™˜ê°’ ê²€ì‚¬ â€” í„°ë¯¸ë„ì— ì¶œë ¥
    if current_app.config['DEBUG_LOGGING_ENABLED']:
        print("DEBUG: get_all_boats() returned", len(registered_boats), "boats")
        for i, b in enumerate(registered_boats, start=1):
            try:
                info = {
                    'repr': repr(b),
                    'type': type(b).__name__,
                    'id': getattr(b, 'id', None),
                    'name': getattr(b, 'name', None),
                    'registered_name': getattr(b, 'registered_name', None),
                    'city': getattr(b, 'city', None),
                    'port': getattr(b, 'port', None),
                    'url': getattr(b, 'url', None),
                }
            except Exception as exc:
                info = {'error': str(exc), 'repr': repr(b)}
            print(f"DEBUG boat[{i}]:", info)

    # ì¡°íšŒ ì‹¤í–‰ - ë³‘ë ¬ ì²˜ë¦¬ë¡œ ì†ë„ ê°œì„ 
    date_str = f"{year:04d}-{month:02d}-{day:02d}"
    results = []
    
    # Flask application contextë¥¼ ìŠ¤ë ˆë“œì—ì„œ ì‚¬ìš©í•˜ê¸° ìœ„í•´ ë¯¸ë¦¬ ì €ì¥
    debug_enabled = current_app.config.get('DEBUG_LOGGING_ENABLED', False)
    
    # ë³‘ë ¬ ì²˜ë¦¬ í•¨ìˆ˜ ì •ì˜
    def process_boat(boat):
        boat_name = getattr(boat, "name", None) or getattr(boat, "registered_name", "unknown")
        boat_url = getattr(boat, "url", "")
        try:
            check = check_single_boat(boat_url, year, month, day, debug_enabled=debug_enabled)
            check_source = check.get("source_url") or boat_url or ""
            
            boat_results = []
            for e in check.get("entries", []):
                full_url = (e.get("used_url") or e.get("source_url") or e.get("url") or check_source or boat_url) or ""
                url_path = e.get("used_url_path") or e.get("url_path") or full_url
                boat_results.append({
                     "registered_name": boat_name,
                     "city": getattr(boat, "city", ""),
                     "port": getattr(boat, "port", ""),
                     "ship_name": e.get("ship_name"),
                     "status": e.get("status"),
                     "available": e.get("available"),
                     "display_status": e.get("display_status"),
                     "raw_status_text": e.get("raw_status_text"),
                     "url": full_url,
                     "url_path": url_path,
                     "fish": e.get("fish"),
                     "row_html": e.get("row_html"),
                     "tide": check.get("tide"),
                })
            return boat_results
        except Exception as e:
            if debug_enabled:
                import traceback
                print(f"Error processing boat {boat_name}: {e}")
                print(traceback.format_exc())
            return []
    
    # ThreadPoolExecutorë¡œ ë³‘ë ¬ ì²˜ë¦¬ (ìµœëŒ€ 10ê°œ ë™ì‹œ ì²˜ë¦¬)
    max_workers = min(10, len(boats_to_query)) if boats_to_query else 1
    
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_boat = {executor.submit(process_boat, boat): boat for boat in boats_to_query}
        for future in as_completed(future_to_boat):
            try:
                boat_results = future.result()
                results.extend(boat_results)
            except Exception as e:
                if debug_enabled:
                    import traceback
                    print(f"Error getting future result: {e}")
                    print(traceback.format_exc())

    # { changed code } : ë“±ë¡ëœ ë°° ëª©ë¡(registered_boats)ì—ì„œ ì§€ì—­ë³„ ë“±ë¡ ìˆ˜ ê³„ì‚°
    region_sets = {}
    for b in get_all_boats():
        city = getattr(b, 'city', None) or ''
        rn = getattr(b, 'name', None) or getattr(b, 'registered_name', None) or ''
        if not city:
            continue
        region_sets.setdefault(city, set()).add(rn or '')
    region_counts = { r: len(s) for r, s in region_sets.items() }
    total_registered = sum(region_counts.values())

    # ì˜ˆì•½ê°€ëŠ¥ ìƒíƒœ ë°°ë¥¼ ë¨¼ì € ë³´ì—¬ì£¼ë„ë¡ ì •ë ¬
    results_sorted = sorted(results, key=lambda x: x.get('status') != 'open')
    # í…œí”Œë¦¿ì—ëŠ” ì‹¤ì œ ë³´ì—¬ì¤„ ê²°ê³¼ ë¦¬ìŠ¤íŠ¸(results)ë¥¼ ì „ë‹¬
    return render_template('status.html',
                           form=form,
                           entries=results_sorted,
                           region_names=region_names,
                           selected_regions=selected_regions,
                           year=year,
                           month=month,
                           day=day,
                           region_counts=region_counts,
                           total_registered=total_registered)

# API endpoint: JSONìœ¼ë¡œ íŒŒì‹±ê²°ê³¼ ë°˜í™˜ (í´ë¼ì´ì–¸íŠ¸ê°€ fetchë¡œ í˜¸ì¶œ)
@views.route('/api/status', methods=['POST'])
def api_status():
    data = request.get_json() or request.form
    try:
        year = int(data.get('year'))
        month = int(data.get('month'))
        day = int(data.get('day'))
    except Exception:
        return jsonify({"error": "invalid date"}), 400

    boats = get_all_boats()
    out = []
    for b in boats:
        info = check_single_boat(b.url, year, month, day, debug_enabled=current_app.config['DEBUG_LOGGING_ENABLED'])
        entries_out = []
        source_url = info.get("source_url") or b.url
        for entry in info.get("entries", []):
            # API ì‘ë‹µì—ì„œë„ ë™ì¼í•œ ìš°ì„ ìˆœìœ„ì™€ ì „ì²´ URL í…ìŠ¤íŠ¸ ì „ë‹¬
            full_url = (entry.get("used_url") or entry.get("source_url") or entry.get("url") or source_url or "") or ""
            url_path = entry.get("used_url_path") or entry.get("url_path") or full_url
            entries_out.append({
                "ship_name": entry.get("ship_name"),
                "status": entry.get("status"),
                "available": entry.get("available"),
                "raw_status_text": entry.get("raw_status_text"),
                "row_html": entry.get("row_html"),
                "source_url": full_url,
                "url_path": url_path,
                "fish": entry.get("fish")
            })
        if not entries_out:
            entries_out.append({
                "ship_name": None,
                "status": "unknown",
                "available": None,
                "raw_status_text": "",
                "row_html": "",
                "source_url": source_url
            })

        out.append({
            "registered_name": b.name,
            "city": b.city,
            "port": b.port,
            "query_date": f"{int(year):04d}-{int(month):02d}-{int(day):02d}",
            "date_id": info.get("date_id"),
            "tide": info.get("tide"),   # ì¶”ê°€: ë¬¼ë•Œ ì •ë³´
            "entries": entries_out
        })
    return jsonify(out)

@views.route('/weather')
def weather():
    """ë‚ ì”¨ ì •ë³´ ì¡°íšŒ í˜ì´ì§€"""
    # map_pageì—ì„œ ì‚¬ìš©í•˜ëŠ” ê²ƒê³¼ ë™ì¼í•œ ë°ì´í„° ì‚¬ìš©
    return render_template('weather.html', 
                         city_port_mapping=get_city_port_mapping(),
                         port_coordinates=get_port_coordinates(),
                         bada_port_ids=get_bada_port_ids())

def get_port_coordinates():
    """í•­êµ¬ ì¢Œí‘œ ì •ë³´ë¥¼ ë°˜í™˜"""
    return {
        'ë‚¨í•­(ì¸ì²œí•­)': {'lat': 37.47, 'lon': 126.62},
        'ì—°ì•ˆë¶€ë‘': {'lat': 37.4416, 'lon': 126.6110},
        'ì˜í¥í•­': {'lat': 37.25455083861362, 'lon': 126.49825493353622},
        'ì˜¤ì´ë„í•­': {'lat': 37.326444939596996, 'lon': 126.65458586308483},
        'ì „ê³¡í•­': {'lat': 37.18786766510414, 'lon': 126.65235743282231},
        'í‰íƒí•­': {'lat': 36.96158755929977, 'lon': 126.84006775074936},
        'ì¥ê³ í•­': {'lat': 37.03122635505709, 'lon': 126.55981703596025},
        'ì‚¼ê¸¸í¬í•­': {'lat': 37.00415509197122, 'lon': 126.45292068915825},
        'ë§ˆê²€í¬í•­': {'lat': 36.61943531903122, 'lon': 126.2875526892295},
        'ëª¨í•­í•­': {'lat': 36.7759, 'lon': 126.1328},
        'ì˜ëª©í•­': {'lat': 36.3999, 'lon': 126.4277},
        'ì‹ ì§„ë„í•­': {'lat': 36.6833, 'lon': 126.1500},
        'ì˜¤ì²œí•­': {'lat': 36.4383319, 'lon': 126.5201303},
        'êµ¬ë§¤í•­': {'lat': 36.424732, 'lon': 126.432133},
        'ëŒ€ì²œí•­': {'lat': 36.3333, 'lon': 126.5167},
        'ë¬´ì°½í¬í•­': {'lat': 36.2436, 'lon': 126.5469},
        'ë‚¨ë‹¹í•­': {'lat': 36.5390947, 'lon': 126.4689945},
        'í™ì›í•­': {'lat': 36.1583, 'lon': 126.5028},
        'ë¹„ì‘í•­': {'lat': 35.93826493213535, 'lon': 126.53099554693064},
        'ì•¼ë¯¸ë„í•­': {'lat': 35.8407672, 'lon': 126.488760},
        'ê²©í¬í•­': {'lat': 35.6225668, 'lon': 126.4694321},
        'ëŒì‚°í•­': {'lat': 34.61326519186631, 'lon': 127.7224984379492},
        'êµ­ë™í•­': {'lat': 34.72949367130133, 'lon': 127.7253480879476},
        'ì†Œí˜¸í•­': {'lat': 34.746193195297266, 'lon': 127.6561636346259},
        'ì‹ ì¶”í•­': {'lat': 34.7308212588099, 'lon': 127.754781729328},
        'ì¢…í¬í•­': {'lat': 34.73738965299665, 'lon': 127.74701532311137},
        'ë…¹ë™ë°©íŒŒì œ': {'lat': 34.52298050694286, 'lon': 127.14353349262528},
    }

def get_city_port_mapping():
    """ì§€ì—­ë³„ í•­êµ¬ ë§¤í•‘ ì •ë³´ë¥¼ ë°˜í™˜"""
    return {
        'ì¸ì²œ': ['ë‚¨í•­(ì¸ì²œí•­)', 'ì—°ì•ˆë¶€ë‘', 'ì˜í¥í•­'],
        'ì•ˆì‚°': ['ì˜¤ì´ë„í•­'],
        'í™”ì„±': ['ì „ê³¡í•­'],
        'í‰íƒ': ['í‰íƒí•­'],
        'ë‹¹ì§„': ['ì¥ê³ í•­'],
        'ì„œì‚°': ['ì‚¼ê¸¸í¬í•­'],
        'íƒœì•ˆ': ['ë§ˆê²€í¬í•­', 'ëª¨í•­í•­', 'ì˜ëª©í•­', 'ì‹ ì§„ë„í•­'],
        'ë³´ë ¹': ['ì˜¤ì²œí•­', 'êµ¬ë§¤í•­', 'ëŒ€ì²œí•­', 'ë¬´ì°½í¬í•­', 'ë‚¨ë‹¹í•­', 'í™ì›í•­'],
        'êµ°ì‚°': ['ë¹„ì‘í•­', 'ì•¼ë¯¸ë„í•­'],
        'ê²©í¬': ['ê²©í¬í•­'],
        'ì—¬ìˆ˜': ['ëŒì‚°í•­', 'êµ­ë™í•­', 'ì†Œí˜¸í•­', 'ì‹ ì¶”í•­', 'ì¢…í¬í•­'],
        'ê³ í¥': ['ë…¹ë™ë°©íŒŒì œ']
    }

def get_bada_port_ids():
    """ë°”ë‹¤íƒ€ì„ í¬íŠ¸ ID ë§¤í•‘ ë°˜í™˜ (í•­êµ¬ëª… -> ID)"""
    return {
        'ë‚¨í•­(ì¸ì²œí•­)': 158,
        'ì—°ì•ˆë¶€ë‘': 158,
        'ì˜í¥í•­': 151,
        'ì˜¤ì´ë„í•­': 380,
        'ì „ê³¡í•­': 618,
        'í‰íƒí•­': 149,
        'ì¥ê³ í•­': 370,
        'ì‚¼ê¸¸í¬í•­': 144,
        'ë§ˆê²€í¬í•­': 1400,
        'ëª¨í•­í•­': 134,
        'ì˜ëª©í•­': 354,
        'ì‹ ì§„ë„í•­': 965,
        'ì˜¤ì²œí•­': 355,
        'êµ¬ë§¤í•­': 1385,
        'ëŒ€ì²œí•­': 126,
        'ë¬´ì°½í¬í•­': 236,
        'ë‚¨ë‹¹í•­': 356,
        'í™ì›í•­': 523,
        'ë¹„ì‘í•­': 118,
        'ì•¼ë¯¸ë„í•­': 348,
        'ê²©í¬í•­': 430,
        'ëŒì‚°í•­': 270,
        'êµ­ë™í•­': 271,
        'ì†Œí˜¸í•­': 826,
        'ì‹ ì¶”í•­': 885,
        'ì¢…í¬í•­': 886,
        'ë…¹ë™ë°©íŒŒì œ': 443,
    }


@views.route('/api/weather', methods=['GET'])
def api_weather():
    """ê¸°ìƒì²­ APIë¥¼ í˜¸ì¶œí•˜ì—¬ ë‚ ì”¨ ì •ë³´ë¥¼ ê°€ì ¸ì˜¤ëŠ” API"""
    import requests
    from datetime import datetime
    
    port = request.args.get('port')
    date_str = request.args.get('date')  # YYYY-MM-DD
    
    if not port or not date_str:
        return jsonify({'error': 'í•­êµ¬ì™€ ë‚ ì§œë¥¼ ì…ë ¥í•´ì£¼ì„¸ìš”.'}), 400
    
    # port_coordinatesì—ì„œ ì¢Œí‘œ ê°€ì ¸ì˜¤ê¸°
    port_coords = get_port_coordinates()
    if port not in port_coords:
        return jsonify({'error': f'{port}ì˜ ì¢Œí‘œ ì •ë³´ë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤.'}), 404
    
    lat = port_coords[port]['lat']
    lon = port_coords[port]['lon']
    
    try:
        # ìœ„ê²½ë„ë¥¼ ê¸°ìƒì²­ ê²©ì ì¢Œí‘œë¡œ ë³€í™˜
        grid = convert_to_grid(lat, lon)
        
        # ë‚ ì§œ íŒŒì‹±
        target_date = datetime.strptime(date_str, '%Y-%m-%d')
        base_date = target_date.strftime('%Y%m%d')
        
        # ê¸°ìƒì²­ ë‹¨ê¸°ì˜ˆë³´ API í˜¸ì¶œ
        # ê³µê³µë°ì´í„°í¬í„¸(https://www.data.go.kr/)ì—ì„œ 'ê¸°ìƒì²­_ë‹¨ê¸°ì˜ˆë³´' ê²€ìƒ‰í•˜ì—¬ API í‚¤ ë°œê¸‰
        service_key = current_app.config.get('KMA_API_KEY', 'd7734746c9c841d53b70df3ffbda3e56422c50e5af2a345ab650bfb24d78b0c9')
        
        # API í‚¤ê°€ ì„¤ì •ë˜ì§€ ì•Šì€ ê²½ìš° í•­êµ¬ë³„ ìƒ˜í”Œ ë°ì´í„° ì‚¬ìš©
        use_sample = (service_key == 'd7734746c9c841d53b70df3ffbda3e56422c50e5af2a345ab650bfb24d78b0c9')
        
        if use_sample:
            # í•­êµ¬ë³„ë¡œ ë‹¤ë¥¸ ìƒ˜í”Œ ë°ì´í„° ë°˜í™˜
            weather_data = generate_sample_weather_data(port, lat, lon)
            return jsonify({
                'lat': lat,
                'lon': lon,
                'nx': grid['nx'],
                'ny': grid['ny'],
                'data': weather_data,
                'note': 'ìƒ˜í”Œ ë°ì´í„°ì…ë‹ˆë‹¤. ì‹¤ì œ ë°ì´í„°ë¥¼ ë³´ë ¤ë©´ ê¸°ìƒì²­ API í‚¤ë¥¼ ì„¤ì •í•´ì£¼ì„¸ìš”.'
            })
        
        # ì‹¤ì œ API í˜¸ì¶œ
        url = 'http://apis.data.go.kr/1360000/VilageFcstInfoService_2.0/getVilageFcst'
        params = {
            'serviceKey': service_key,
            'pageNo': '1',
            'numOfRows': '1000',
            'dataType': 'JSON',
            'base_date': base_date,
            'base_time': '0500',
            'nx': grid['nx'],
            'ny': grid['ny']
        }
        
        response = requests.get(url, params=params, timeout=10)
        
        if response.status_code != 200:
            # API í˜¸ì¶œ ì‹¤íŒ¨ ì‹œ ìƒ˜í”Œ ë°ì´í„°ë¡œ ëŒ€ì²´
            current_app.logger.warning(f"KMA API call failed with status {response.status_code}")
            weather_data = generate_sample_weather_data(port, lat, lon)
            return jsonify({
                'lat': lat,
                'lon': lon,
                'nx': grid['nx'],
                'ny': grid['ny'],
                'data': weather_data,
                'note': 'API í˜¸ì¶œ ì‹¤íŒ¨ë¡œ ìƒ˜í”Œ ë°ì´í„°ë¥¼ í‘œì‹œí•©ë‹ˆë‹¤.'
            })
        
        result = response.json()
        
        # API ì‘ë‹µ ì²˜ë¦¬
        weather_data = process_kma_weather_data(result, base_date)
        
        if not weather_data:
            # ë°ì´í„°ê°€ ì—†ëŠ” ê²½ìš° ìƒ˜í”Œ ë°ì´í„°ë¡œ ëŒ€ì²´
            weather_data = generate_sample_weather_data(port, lat, lon)
            return jsonify({
                'lat': lat,
                'lon': lon,
                'nx': grid['nx'],
                'ny': grid['ny'],
                'data': weather_data,
                'note': 'í•´ë‹¹ ë‚ ì§œì˜ ì‹¤ì œ ë°ì´í„°ê°€ ì—†ì–´ ìƒ˜í”Œ ë°ì´í„°ë¥¼ í‘œì‹œí•©ë‹ˆë‹¤.'
            })
        
        return jsonify({
            'lat': lat,
            'lon': lon,
            'nx': grid['nx'],
            'ny': grid['ny'],
            'data': weather_data
        })
        
    except Exception as e:
        current_app.logger.error(f"Weather API error: {e}")
        # ì—ëŸ¬ ë°œìƒ ì‹œì—ë„ ìƒ˜í”Œ ë°ì´í„° ì œê³µ
        try:
            weather_data = generate_sample_weather_data(port, lat, lon)
            return jsonify({
                'lat': lat,
                'lon': lon,
                'nx': grid['nx'] if 'grid' in locals() else 0,
                'ny': grid['ny'] if 'grid' in locals() else 0,
                'data': weather_data,
                'error': f'ì—ëŸ¬ê°€ ë°œìƒí•˜ì—¬ ìƒ˜í”Œ ë°ì´í„°ë¥¼ í‘œì‹œí•©ë‹ˆë‹¤: {str(e)}'
            })
        except:
            return jsonify({'error': f'ë‚ ì”¨ ì •ë³´ë¥¼ ê°€ì ¸ì˜¬ ìˆ˜ ì—†ìŠµë‹ˆë‹¤: {str(e)}'}), 500

def convert_to_grid(lat, lon):
    """ìœ„ê²½ë„ë¥¼ ê¸°ìƒì²­ ê²©ì ì¢Œí‘œë¡œ ë³€í™˜"""
    import math
    
    RE = 6371.00877  # ì§€êµ¬ ë°˜ê²½(km)
    GRID = 5.0  # ê²©ì ê°„ê²©(km)
    SLAT1 = 30.0  # í‘œì¤€ìœ„ë„1
    SLAT2 = 60.0  # í‘œì¤€ìœ„ë„2
    OLON = 126.0  # ê¸°ì¤€ì  ê²½ë„
    OLAT = 38.0  # ê¸°ì¤€ì  ìœ„ë„
    XO = 43  # ê¸°ì¤€ì  Xì¢Œí‘œ
    YO = 136  # ê¸°ì¤€ì  Yì¢Œí‘œ

    DEGRAD = math.pi / 180.0
    re = RE / GRID
    slat1 = SLAT1 * DEGRAD
    slat2 = SLAT2 * DEGRAD
    olon = OLON * DEGRAD
    olat = OLAT * DEGRAD

    sn = math.tan(math.pi * 0.25 + slat2 * 0.5) / math.tan(math.pi * 0.25 + slat1 * 0.5)
    sn = math.log(math.cos(slat1) / math.cos(slat2)) / math.log(sn)
    sf = math.tan(math.pi * 0.25 + slat1 * 0.5)
    sf = math.pow(sf, sn) * math.cos(slat1) / sn
    ro = math.tan(math.pi * 0.25 + olat * 0.5)
    ro = re * sf / math.pow(ro, sn)

    ra = math.tan(math.pi * 0.25 + lat * DEGRAD * 0.5)
    ra = re * sf / math.pow(ra, sn)
    theta = lon * DEGRAD - olon
    if theta > math.pi:
        theta -= 2.0 * math.pi
    if theta < -math.pi:
        theta += 2.0 * math.pi
    theta *= sn

    nx = int(ra * math.sin(theta) + XO + 0.5)
    ny = int(ro - ra * math.cos(theta) + YO + 0.5)

    return {'nx': nx, 'ny': ny}

def process_kma_weather_data(result, base_date):
    """ê¸°ìƒì²­ API ì‘ë‹µ ë°ì´í„° ì²˜ë¦¬"""
    if not result.get('response', {}).get('body', {}).get('items', {}).get('item'):
        return []
    
    items = result['response']['body']['items']['item']
    time_data = {}
    
    # ì‹œê°„ëŒ€ë³„ë¡œ ë°ì´í„° ê·¸ë£¹í™”
    for item in items:
        fcst_date = item['fcstDate']
        fcst_time = item['fcstTime']
        category = item['category']
        value = item['fcstValue']
        
        if fcst_date == base_date:
            time_key = f"{fcst_time[:2]}ì‹œ"
            if time_key not in time_data:
                time_data[time_key] = {}
            time_data[time_key][category] = value
    
    # ì‹œê°„ëŒ€ë³„ ë°ì´í„°ë¥¼ ë°°ì—´ë¡œ ë³€í™˜
    weather_array = []
    for time in sorted(time_data.keys()):
        data = time_data[time]
        
        # í’í–¥ ë³€í™˜
        wind_dir_deg = float(data.get('VEC', 0))
        direction = get_wind_direction(wind_dir_deg)
        
        # ë‚ ì”¨ ì•„ì´ì½˜ ê²°ì •
        sky = data.get('SKY', '1')
        pty = data.get('PTY', '0')
        weather = get_weather_icon(sky, pty)
        
        weather_array.append({
            'time': time,
            'direction': direction,
            'windSpeed': float(data.get('WSD', 0)),
            'maxWindSpeed': float(data.get('WSD', 0)) * 1.5,
            'weather': weather,
            'temp': float(data.get('TMP', 0)),
            'waveHeight': 0.6,  # ê¸°ë³¸ê°’
            'wavePeriod': 7.0   # ê¸°ë³¸ê°’
        })
    
    return weather_array

def get_wind_direction(deg):
    """í’í–¥ ê°ë„ë¥¼ ë°©ìœ„ë¡œ ë³€í™˜"""
    dirs = ['ë¶', 'ë¶ë¶ë™', 'ë¶ë™', 'ë™ë¶ë™', 'ë™', 'ë™ë‚¨ë™', 'ë‚¨ë™', 'ë‚¨ë‚¨ë™',
            'ë‚¨', 'ë‚¨ë‚¨ì„œ', 'ë‚¨ì„œ', 'ì„œë‚¨ì„œ', 'ì„œ', 'ì„œë¶ì„œ', 'ë¶ì„œ', 'ë¶ë¶ì„œ']
    idx = int((deg + 22.5 * 0.5) / 22.5) % 16
    return dirs[idx]

def get_weather_icon(sky, pty):
    """í•˜ëŠ˜ ìƒíƒœì™€ ê°•ìˆ˜ í˜•íƒœë¡œ ë‚ ì”¨ ì•„ì´ì½˜ ê²°ì •"""
    if pty == '1' or pty == '4':
        return 'ğŸŒ§ï¸'  # ë¹„
    if pty == '2':
        return 'ğŸŒ¨ï¸'  # ë¹„/ëˆˆ
    if pty == '3':
        return 'â„ï¸'  # ëˆˆ
    if sky == '1':
        return 'â˜€ï¸'  # ë§‘ìŒ
    if sky == '3':
        return 'â›…'  # êµ¬ë¦„ë§ìŒ
    if sky == '4':
        return 'â˜ï¸'  # íë¦¼
    return 'ğŸŒ¤ï¸'

def generate_sample_weather_data(port_name, lat, lon):
    """í•­êµ¬ë³„ë¡œ ë‹¤ë¥¸ ìƒ˜í”Œ ë‚ ì”¨ ë°ì´í„° ìƒì„±"""
    import random
    
    # í•­êµ¬ ì´ë¦„ì„ ì‹œë“œë¡œ ì‚¬ìš©í•˜ì—¬ ì¼ê´€ëœ ëœë¤ ê°’ ìƒì„±
    seed = hash(port_name) % 10000
    random.seed(seed)
    
    times = ['00ì‹œ', '03ì‹œ', '06ì‹œ', '09ì‹œ', '12ì‹œ', '15ì‹œ', '18ì‹œ', '21ì‹œ']
    
    # ìœ„ë„ì— ë”°ë¼ ê¸°ì˜¨ ë²”ìœ„ ì¡°ì • (ë‚¨ìª½ì´ ë” ë”°ëœ»í•¨)
    base_temp = 15 + (37.5 - lat) * 0.5  # ìœ„ë„ê°€ ë‚®ì„ìˆ˜ë¡ ê¸°ì˜¨ ë†’ìŒ
    
    # ê²½ë„ì™€ ìœ„ë„ë¡œ í’í–¥ ê²½í–¥ ê²°ì •
    wind_direction_base = int((lon - 126) * 10 + (lat - 35) * 5) % 360
    
    data = []
    for i, time in enumerate(times):
        # ì‹œê°„ëŒ€ë³„ ê¸°ì˜¨ ë³€í™”
        hour = int(time.replace('ì‹œ', ''))
        temp_variation = -3 if hour < 6 else (5 if 12 <= hour < 15 else 0)
        temp = round(base_temp + temp_variation + random.uniform(-2, 2), 1)
        
        # í’í–¥ (í•­êµ¬ë³„ë¡œ ë‹¤ë¥´ê²Œ)
        wind_deg = (wind_direction_base + random.randint(-30, 30)) % 360
        direction = get_wind_direction(wind_deg)
        
        # í’ì† (ì—°ì•ˆ ì§€ì—­ íŠ¹ì„±)
        wind_speed = round(random.uniform(1.5, 6.0), 1)
        max_wind_speed = round(wind_speed * random.uniform(1.3, 1.8), 1)
        
        # ë‚ ì”¨ (ì¼ë¶€ ëœë¤)
        weather_options = ['â˜€ï¸', 'ğŸŒ¤ï¸', 'â›…', 'â˜ï¸']
        if random.random() < 0.15:  # 15% í™•ë¥ ë¡œ ë¹„
            weather_options = ['ğŸŒ§ï¸', 'ğŸŒ¦ï¸']
        weather = random.choice(weather_options)
        
        # íŒŒê³  (í’ì†ê³¼ ì—°ê´€)
        wave_height = round(wind_speed * 0.15 + random.uniform(0.3, 0.8), 1)
        wave_period = round(random.uniform(4.0, 9.0), 1)
        
        data.append({
            'time': time,
            'direction': direction,
            'windSpeed': wind_speed,
            'maxWindSpeed': max_wind_speed,
            'weather': weather,
            'temp': int(temp),
            'waveHeight': wave_height,
            'wavePeriod': wave_period
        })
    
    return data

# ---------------- Tide (Badatime) Integration -----------------
@views.route('/api/tide')
def api_tide():
    """ë°”ë‹¤íƒ€ì„ íŠ¹ì • í•­êµ¬ ë²ˆí˜¸(port_id)ì˜ ì£¼ê°„(week_container) ì •ë³´ë¥¼ íŒŒì‹±í•˜ì—¬ ì‹œê°„ëŒ€ë³„ ë°ì´í„° ë°˜í™˜.
    ìš”ì²­: /api/tide?port_id=118
    ë°˜í™˜ í•„ë“œ: time, wind_dir, wind_speed, weather, temperature, wave_info
    ë°”ë‹¤íƒ€ì„ í˜ì´ì§€ì— í’í–¥/í’ì†/ë‚ ì”¨/ê¸°ì˜¨/íŒŒê³ ê°€ ëª¨ë‘ ì—†ì„ ìˆ˜ ìˆìœ¼ë¯€ë¡œ ê°€ìš©í•œ ì •ë³´ë§Œ êµ¬ì„±í•˜ê³  ë‚˜ë¨¸ì§€ëŠ” ì¶”ì •/ë¹ˆê°’ ì²˜ë¦¬.
    """
    import requests
    from bs4 import BeautifulSoup
    port_id = request.args.get('port_id', type=int)
    if not port_id:
        return jsonify({'error': 'port_id íŒŒë¼ë¯¸í„°ê°€ í•„ìš”í•©ë‹ˆë‹¤.'}), 400

    # ë‚ ì§œëŠ” /{port_id}/tide/YYYY-MM-DD í˜•íƒœì˜ ê²½ë¡œë¡œ ì „ë‹¬ë¨
    date_str = request.args.get('date')  # YYYY-MM-DD
    base_url = f"https://www.badatime.com/{port_id}/tide"
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0 Safari/537.36'
    }
    try:
        # ë‚ ì§œê°€ ìˆìœ¼ë©´ ê²½ë¡œ ì„¸ê·¸ë¨¼íŠ¸ë¡œ ì „ë‹¬: /{port}/tide/YYYY-MM-DD
        used_url = f"{base_url}/{date_str}" if date_str else base_url
        resp = requests.get(used_url, headers=headers, timeout=10)
        if resp.status_code != 200:
            return jsonify({'error': f'í˜ì´ì§€ ì‘ë‹µ ì˜¤ë¥˜: {resp.status_code}'}), 502
    except Exception as e:
        return jsonify({'error': f'ìš”ì²­ ì‹¤íŒ¨: {e}'}), 500

    soup = BeautifulSoup(resp.text, 'html.parser')
    week_container = soup.select_one('.week_container')
    if not week_container:
        return jsonify({'error': 'week_container(class)ë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤.'}), 500

    table = week_container.select_one('table.week_table')
    if not table:
        return jsonify({'error': 'week_tableì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤.'}), 500

    import re
    rows = table.select('tbody > tr')
    if not rows or len(rows) < 5:
        return jsonify({'error': 'ì˜ˆìƒë³´ë‹¤ ì ì€ í–‰. êµ¬ì¡° ë³€ê²½ ê°€ëŠ¥ì„±.'}), 500

    # 1í–‰: ë‚ ì§œ + ì‹œê°„ í—¤ë”ë“¤
    time_cells = rows[0].find_all('td')[1:]  # ì²«ë²ˆì§¸ëŠ” ë‚ ì§œ
    times = []
    for c in time_cells:
        t = c.get_text(strip=True).replace('í˜„ì¬','').strip()
        # Normalize '07ì‹œ' -> '07ì‹œ'
        times.append(t)

    count = len(times)

    def extract_icon_row(tr):
        icons = []
        for td in tr.find_all('td')[1:]:
            img = td.find('img')
            icons.append(img['src'] if img else '')
        return icons

    def extract_text_cells(tr):
        return [td.get_text(strip=True) for td in tr.find_all('td')[1:]]

    # í–‰ ì‹ë³„: ë‘ë²ˆì§¸ í–‰ ì•„ì´ì½˜, ì„¸ë²ˆì§¸ í–‰ ë‚ ì”¨í…ìŠ¤íŠ¸(ë§‘ìŒ), ë„¤ë²ˆì§¸ ê¸°ì˜¨(ì²«ì…€ 'ê¸°ì˜¨'), ë‹¤ì„¯ë²ˆì§¸ í’í–¥(ì²«ì…€ 'í’í–¥'), ì—¬ì„¯ë²ˆì§¸ í’ì†, ì¼ê³±ë²ˆì§¸ íŒŒê³ , ì—¬ëŸë²ˆì§¸ ìŠµë„, ì•„í™‰ë²ˆì§¸ ê°•ìˆ˜ëŸ‰
    icon_urls      = extract_icon_row(rows[1])
    weather_texts  = extract_text_cells(rows[2])
    temp_values    = extract_text_cells(rows[3]) if 'ê¸°ì˜¨' in rows[3].find('td').get_text() else ['']*count
    wind_dir_cells = rows[4].find_all('td')[1:]
    wind_dirs = []
    wind_dir_icons = []
    for td in wind_dir_cells:
        img = td.find('img')
        wind_dir_icons.append(img['src'] if img else '')
        # span ë˜ëŠ” í…ìŠ¤íŠ¸
        txt = td.get_text(strip=True)
        wind_dirs.append(txt)
    wind_speeds    = extract_text_cells(rows[5]) if 'í’ì†' in rows[5].find('td').get_text() else ['']*count
    wave_heights   = extract_text_cells(rows[6]) if 'íŒŒê³ ' in rows[6].find('td').get_text() else ['']*count
    humidities     = extract_text_cells(rows[7]) if len(rows) > 7 and 'ìŠµë„' in rows[7].find('td').get_text() else ['']*count
    precipitations = extract_text_cells(rows[8]) if len(rows) > 8 and 'ê°•ìˆ˜' in rows[8].find('td').get_text() else ['']*count

    data_out = []
    for i in range(count):
        data_out.append({
            'time': times[i],
            'weather_icon_url': icon_urls[i] if i < len(icon_urls) else '',
            'weather_text': weather_texts[i] if i < len(weather_texts) else '',
            'temperature': temp_values[i] if i < len(temp_values) else '',
            'wind_dir': wind_dirs[i] if i < len(wind_dirs) else '',
            'wind_dir_icon_url': wind_dir_icons[i] if i < len(wind_dir_icons) else '',
            'wind_speed': wind_speeds[i] if i < len(wind_speeds) else '',
            'wave_height': wave_heights[i] if i < len(wave_heights) else '',
            'humidity': humidities[i] if i < len(humidities) else '',
            'precipitation': precipitations[i] if i < len(precipitations) else ''
        })

    return jsonify({'port_id': port_id, 'source_url': used_url if date_str else base_url, 'data': data_out, 'date': date_str})

# New: Parse Badatime graph page and return only summary table + chart script
@views.route('/api/tide_graph', methods=['GET'])
def api_tide_graph():
    """Badatime ê·¸ë˜í”„ í˜ì´ì§€(/{port_id}/graph/{date})ì—ì„œ ìš”ì•½ í…Œì´ë¸”(pc_txt_view)ê³¼
    ì°¨íŠ¸ ì»¨í…Œì´ë„ˆ(#chartdiv) ë° í•´ë‹¹ ìŠ¤í¬ë¦½íŠ¸ë§Œ ì¶”ì¶œí•´ì„œ ë°˜í™˜.
    ì‘ë‹µ: { success, pc_html, chart_html, script, source_url }
    """
    import requests
    from bs4 import BeautifulSoup

    port_id = request.args.get('port_id', type=int)
    date_str = request.args.get('date', default='')  # YYYY-MM-DD
    if not port_id or not date_str:
        return jsonify({'success': False, 'message': 'port_idì™€ dateê°€ í•„ìš”í•©ë‹ˆë‹¤.'}), 400

    source_url = f"https://www.badatime.com/{port_id}/graph/{date_str}"
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0 Safari/537.36'
    }
    try:
        resp = requests.get(source_url, headers=headers, timeout=10)
        if resp.status_code != 200:
            return jsonify({'success': False, 'message': f'í˜ì´ì§€ ì‘ë‹µ ì˜¤ë¥˜: {resp.status_code}'}), 502
    except Exception as e:
        return jsonify({'success': False, 'message': f'ìš”ì²­ ì‹¤íŒ¨: {e}'}), 500

    soup = BeautifulSoup(resp.text, 'html.parser')

    # PC ìš”ì•½ í…Œì´ë¸”
    pc_view = soup.select_one('div.pc_txt_view')
    pc_html = pc_view.decode() if pc_view else ''

    # ì°¨íŠ¸ ì»¨í…Œì´ë„ˆì™€ ìŠ¤í¬ë¦½íŠ¸(ë°”ë¡œ ë’¤ì— ì˜¤ëŠ” inline script)
    chart_div = soup.select_one('#chartdiv')
    chart_html = ''
    script_text = ''
    if chart_div:
        # chart div ìì²´ëŠ” ë³´í†µ ë¹ˆ div. height ìŠ¤íƒ€ì¼ì„ ë³´ì¥í•˜ê¸° ìœ„í•´ ê¸°ë³¸ ë†’ì´ ë¶€ì—¬
        # ì›ë³¸ divë¥¼ ë³µì‚¬í•˜ê³  style ì¶”ê°€
        chart_div_copy = BeautifulSoup(str(chart_div), 'html.parser')
        chart_root = chart_div_copy.select_one('#chartdiv')
        if chart_root:
            # ê¸°ë³¸ ë†’ì´ ì ìš© (ì—†ì„ ê²½ìš°)
            style_val = chart_root.get('style', '')
            if 'height:' not in style_val:
                style_val = (style_val + '; height: 460px;').strip('; ')
                chart_root['style'] = style_val
        chart_html = str(chart_div_copy)

        # ì°¨íŠ¸ ì„¤ì • ìŠ¤í¬ë¦½íŠ¸: chartdiv ë‹¤ìŒ <script> ì¶”ì¶œ
        next_script = chart_div.find_next('script')
        if next_script and next_script.string:
            script_text = next_script.string
        else:
            # ì¼ë¶€ í˜ì´ì§€ëŠ” script ë‚´ì— ì£¼ì„/ê³µë°± í¬í•¨ -> ì „ì²´ í…ìŠ¤íŠ¸ ì‚¬ìš©
            script_text = next_script.get_text("\n") if next_script else ''

        # ì•ˆì „ì„ ìœ„í•´ ì™¸ë¶€ ì°¸ì¡°ê°€ ìƒëŒ€ê²½ë¡œì¼ ê²½ìš° ì ˆëŒ€ê²½ë¡œë¡œ ê³ ì¹˜ê¸°(ì´ë¯¸ì§€/ì•„ì´ì½˜ ë“±)
        def absolutize_urls(html_text: str) -> str:
            return re.sub(r'(["\'])(\/\/(?:images|img)\.badatime\.com[^"\']*)(["\'])', r"http:\1\2\3", html_text)

        pc_html = absolutize_urls(pc_html)
        chart_html = absolutize_urls(chart_html)

    return jsonify({
        'success': True,
        'pc_html': pc_html,
        'chart_html': chart_html,
        'script': script_text,
        'source_url': source_url,
    })

@views.route('/map')
def map_page():
    """ì§€ë„ í˜ì´ì§€ - í•­êµ¬ë³„ ë“±ë¡ëœ ë°° í‘œì‹œ"""
    port_coordinates = get_port_coordinates()
    city_port_mapping = get_city_port_mapping()

    boats = get_all_boats()
    boat_counts = {}
    port_boat_names = {}
    for boat in boats:
        port = boat.port
        if port not in port_boat_names:
            port_boat_names[port] = []
        port_boat_names[port].append(boat.name)

    for boat in boats:
        port = boat.port
        if port in boat_counts:
            boat_counts[port] += 1
        else:
            boat_counts[port] = 1

    total_boats = len(boats)

    return render_template(
        'map.html',
        city_port_mapping=city_port_mapping,
        port_coordinates=port_coordinates,
        boat_counts=boat_counts,
        port_boat_names=port_boat_names,
        total_boats=total_boats
    )

# ì¶”ê°€: ë°° ì‚­ì œ ë¼ìš°íŠ¸ (POST)
@views.route('/delete/<int:boat_id>', methods=['POST'], endpoint='delete_boat')
def delete_boat_route(boat_id):
    try:
        delete_boat(boat_id)
        flash('ë°°ê°€ ì‚­ì œë˜ì—ˆìŠµë‹ˆë‹¤.', 'success')
    except Exception as e:
        flash(f'ì‚­ì œ ì¤‘ ì˜¤ë¥˜: {e}', 'danger')
    return redirect(url_for('views.index'))

# New route: handle deletion of selected boats
@views.route('/delete_boats', methods=['POST'])
def delete_boats():
    ids = request.form.getlist('delete_ids')
    if not ids:
        flash('ì‚­ì œí•  ë°°ë¥¼ ì„ íƒí•˜ì„¸ìš”.', 'warning')
        return redirect(url_for('views.index'))
    deleted = 0
    for bid in ids:
        try:
            # delete_boat í•¨ìˆ˜ê°€ idë¥¼ ë°›ëŠ”ë‹¤ê³  ê°€ì •
            delete_boat(int(bid))
            deleted += 1
        except Exception as e:
            # continue on error, but notify
            print(f"delete_boat error for id={bid}: {e}")
    flash(f'{deleted}ê°œì˜ ë°°ê°€ ì‚­ì œë˜ì—ˆìŠµë‹ˆë‹¤.', 'success')
    return redirect(url_for('views.index'))

@views.route('/upload_excel', methods=['POST'])
def upload_excel():
    from models import Boat
    from db import db

    if 'excel_file' not in request.files:
        return jsonify({'success': False, 'message': 'íŒŒì¼ì´ ì—†ìŠµë‹ˆë‹¤.'}), 400
    
    file = request.files['excel_file']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'íŒŒì¼ì„ ì„ íƒí•´ì£¼ì„¸ìš”.'}), 400

    if file and (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        try:
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            # Get existing boat names to check for duplicates
            existing_names = {b.name for b in Boat.query.all()}
            
            new_boats_count = 0
            # Iterate over rows, skipping the header (row 1)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Column order from download_excel: No, ì§€ì—­, í•­êµ¬, ë“±ë¡ëœ ë°°, URL
                # We ignore 'No' (index 0)
                if len(row) < 5:
                    continue # Skip malformed rows

                city = row[1]
                port = row[2]
                name = row[3]
                url = row[4]

                # Basic validation
                if not all([city, port, name, url]):
                    current_app.logger.warning(f"Skipping row due to missing data: {row}")
                    continue

                if name not in existing_names:
                    new_boat = Boat(name=name, url=url, city=city, port=port)
                    db.session.add(new_boat)
                    existing_names.add(name) # Avoid duplicates from within the file
                    new_boats_count += 1
            
            db.session.commit()
            
            if new_boats_count > 0:
                message = f'ì„±ê³µ: {new_boats_count}ì²™ì˜ ìƒˆë¡œìš´ ë°°ë¥¼ ë“±ë¡í–ˆìŠµë‹ˆë‹¤.'
            else:
                message = 'ì¶”ê°€í•  ìƒˆë¡œìš´ ë°°ê°€ ì—†ìŠµë‹ˆë‹¤. ëª¨ë“  ë°°ê°€ ì´ë¯¸ ë“±ë¡ë˜ì–´ ìˆìŠµë‹ˆë‹¤.'

            return jsonify({'success': True, 'message': message})

        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Excel upload failed: {e}")
            return jsonify({'success': False, 'message': f'íŒŒì¼ ì²˜ë¦¬ ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤: {e}'}), 500

    return jsonify({'success': False, 'message': 'ì—‘ì…€ íŒŒì¼(.xlsx, .xls)ë§Œ ì—…ë¡œë“œí•  ìˆ˜ ìˆìŠµë‹ˆë‹¤.'}), 400

# API ì—”ë“œí¬ì¸íŠ¸: ì„ ë°• ëª©ë¡ JSONìœ¼ë¡œ ë°˜í™˜
@views.route('/api/ships', methods=['GET'])
def api_ships():
    """ì„ ë°• ëª©ë¡ì„ JSON í˜•íƒœë¡œ ë°˜í™˜í•˜ëŠ” API ì—”ë“œí¬ì¸íŠ¸"""
    try:
        boats = get_all_boats()
        ships_data = []
        
        for boat in boats:
            ship = {
                'id': boat.id,
                'region': boat.city,  # ì§€ì—­
                'port': boat.port,    # í•­êµ¬
                'registration_number': boat.name,  # ë“±ë¡ë²ˆí˜¸ (í˜„ì¬ëŠ” nameì„ ì‚¬ìš©)
                'name': boat.name,    # ì„ ë°• ì´ë¦„
                'url': boat.url       # ìƒì„¸ URL
            }
            ships_data.append(ship)
        
        return jsonify(ships_data)
    
    except Exception as e:
        current_app.logger.error(f"API ships error: {e}")
        return jsonify({'error': 'ì„ ë°• ëª©ë¡ì„ ê°€ì ¸ì˜¤ëŠ” ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤.'}), 500

# API ì—”ë“œí¬ì¸íŠ¸: ìƒˆ ì„ ë°• ë“±ë¡
@views.route('/api/ships', methods=['POST'])
def api_add_ship():
    """ìƒˆ ì„ ë°•ì„ ë“±ë¡í•˜ëŠ” API ì—”ë“œí¬ì¸íŠ¸"""
    try:
        data = request.get_json()
        
        # í•„ìˆ˜ í•„ë“œ ê²€ì¦
        required_fields = ['region', 'port', 'registrationNumber', 'url']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'error': f'{field} í•„ë“œê°€ í•„ìš”í•©ë‹ˆë‹¤.'}), 400
        
        # ì„ ë°• ë“±ë¡
        add_boat_instance(
            name=data.get('registrationNumber'),  # ë“±ë¡ë²ˆí˜¸ë¥¼ nameìœ¼ë¡œ ì‚¬ìš©
            url=data.get('url'),
            city=data.get('region'),
            port=data.get('port')
        )
        
        return jsonify({'success': True, 'message': 'ì„ ë°•ì´ ì„±ê³µì ìœ¼ë¡œ ë“±ë¡ë˜ì—ˆìŠµë‹ˆë‹¤.'})
        
    except Exception as e:
        current_app.logger.error(f"API add ship error: {e}")
        return jsonify({'error': 'ì„ ë°• ë“±ë¡ ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤.'}), 500

